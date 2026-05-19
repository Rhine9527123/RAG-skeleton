from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date, timedelta
import random

wb = Workbook()
ws = wb.active
ws.title = "sales_detail"

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="2F5496")
money_fmt = "#,##0.00"
pct_fmt = "0.0%"
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
center = Alignment(horizontal="center", vertical="center")

headers = ["date", "weekday", "weather", "customers", "revenue", "food_cost", "drinks", "profit_rate", "note"]
col_names = ["\u65e5\u671f", "\u661f\u671f", "\u5929\u6c14", "\u5ba2\u6d41(\u4eba)", "\u8425\u4e1a\u989d(\u5143)", "\u98df\u6750\u6210\u672c(\u5143)", "\u9152\u6c34\u6536\u5165(\u5143)", "\u6bdb\u5229\u7387", "\u5907\u6ce8"]
for col, h in enumerate(col_names, 1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = thin_border

weathers = ["sunny 32C", "sunny 34C", "cloudy 30C", "showers 28C", "heavy rain 25C",
            "overcast 26C", "sunny 33C", "sunny 36C", "thunderstorm 27C", "cloudy 29C",
            "sunny 31C", "overcast 27C", "light rain 24C", "sunny 35C", "sunny 37C"]
weather_cn = ["晴天32度", "晴天34度", "多云30度", "阵雨28度", "大雨25度",
              "阴天26度", "晴天33度", "晴天36度", "雷阵雨27度", "多云29度",
              "晴天31度", "阴天27度", "小雨24度", "晴天35度", "晴天37度"]
weekdays_cn = ["周一", "周二", "周三", "周四", "周五", "周六", "周日"]

base_date = date(2026, 4, 2)
random.seed(42)

for i in range(15):
    row = i + 2
    d = base_date + timedelta(days=i)
    wd = weekdays_cn[d.weekday()]
    is_weekend = d.weekday() >= 5
    weather = weather_cn[i]

    rain = "雨" in weather
    hot = any(x in weather for x in ["34", "35", "36", "37"])
    base_customers = random.randint(80, 130) if is_weekend else random.randint(50, 85)
    if rain:
        base_customers = int(base_customers * 0.6)
    if hot:
        base_customers = int(base_customers * 1.2)
    customers = base_customers

    avg_per = random.randint(25, 45)
    revenue = customers * avg_per + random.randint(-200, 300)
    food_cost = int(revenue * random.uniform(0.35, 0.42))
    drinks = int(revenue * random.uniform(0.15, 0.28))

    notes = ""
    if is_weekend:
        notes = "weekend peak"
    if rain:
        notes += " rain low traffic"
    if hot:
        notes += " hot weather drinks sell well"
    if i == 13:
        notes = "beer promotion"
    if i == 6:
        notes = "construction team party"

    note_cn = ""
    if is_weekend:
        note_cn = "weekend peak"
    if rain:
        note_cn += " rain traffic down"
    if hot:
        note_cn += " hot drinks good"
    if i == 13:
        note_cn = "beer promotion event"
    if i == 6:
        note_cn = "nearby construction team booking"

    ws.cell(row=row, column=1, value=d.strftime("%Y-%m-%d")).alignment = center
    ws.cell(row=row, column=2, value=wd).alignment = center
    ws.cell(row=row, column=3, value=weather).alignment = center
    ws.cell(row=row, column=4, value=customers).alignment = center
    ws.cell(row=row, column=5, value=revenue).number_format = money_fmt
    ws.cell(row=row, column=6, value=food_cost).number_format = money_fmt
    ws.cell(row=row, column=7, value=drinks).number_format = money_fmt
    ws.cell(row=row, column=8).value = "=1-(F{}/E{})".format(row, row)
    ws.cell(row=row, column=8).number_format = pct_fmt
    ws.cell(row=row, column=8).alignment = center
    ws.cell(row=row, column=9, value=note_cn.strip()).alignment = center

    for col in range(1, 10):
        ws.cell(row=row, column=col).border = thin_border

total_row = 17
ws.cell(row=total_row, column=1, value="合计").font = Font(bold=True)
ws.cell(row=total_row, column=4, value="=SUM(D2:D16)")
ws.cell(row=total_row, column=5, value="=SUM(E2:E16)")
ws.cell(row=total_row, column=6, value="=SUM(F2:F16)")
ws.cell(row=total_row, column=7, value="=SUM(G2:G16)")
ws.cell(row=total_row, column=8, value="=1-(F{}/E{})".format(total_row, total_row))
ws.cell(row=total_row, column=8).number_format = pct_fmt
for col in range(1, 10):
    c = ws.cell(row=total_row, column=col)
    c.font = Font(bold=True)
    c.border = thin_border
    c.alignment = center
    if col in [5, 6, 7]:
        c.number_format = money_fmt

widths = [14, 8, 14, 10, 14, 14, 14, 10, 28]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[chr(64 + i)].width = w

wb.save("./data/sales_data.xlsx")
print("OK")
